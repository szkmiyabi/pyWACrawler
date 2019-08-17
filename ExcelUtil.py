from openpyxl import Workbook
from openpyxl.styles import PatternFill, Border, Side, Font, Alignment

class ExcelUtil:

    # 最大文字数32767に収める
    @classmethod
    def fetch_overflow_characters(cls, data):
        if len(data) >= 32767:
            prefix = "【注意】セルに入力可能な文字数の上限を超えました。32767文字以降は切り捨てられます。\n\n"
            return prefix + data[:32700]
        else:
            return data

    # Excelファイルに出力
    @staticmethod
    def save_xlsx(datas):
        wb = Workbook()
        ws = wb.active
        border_style = Border(
            left = Side(border_style="thin", color='FF000000'),
            right = Side(border_style="thin", color='FF000000'),
            top = Side(border_style="thin", color='FF000000'),
            bottom = Side(border_style="thin", color='FF000000'),
        )
        r = 1
        for parent_row in datas:
            c = 1
            for child_row in parent_row:
                col = ExcelUtil.fetch_overflow_characters(child_row)
                ws.cell(row=r, column=c).value = col
                ws.cell(row=r, column=c).border = border_style
                ws.cell(row=r, column=c).alignment = Alignment(vertical="top")
                c += 1
            cellobj = ws.cell(row=r, column=6)
            if cellobj.value == "適合":
                for i in range(1, len(parent_row)+1):
                    ws.cell(row=r, column=i).fill = PatternFill(fill_type="solid", fgColor="FF40FFFF")
            elif cellobj.value == "適合(注記)":
                for i in range(1, len(parent_row)+1):
                    ws.cell(row=r, column=i).fill = PatternFill(fill_type="solid", fgColor="FF40FF40")
            elif cellobj.value == "不適合":
                for i in range(1, len(parent_row)+1):
                    ws.cell(row=r, column=i).fill = PatternFill(fill_type="solid", fgColor="FFFF8080")
            elif cellobj.value == "非適用":
                for i in range(1, len(parent_row)+1):
                    ws.cell(row=r, column=i).fill = PatternFill(fill_type="solid", fgColor="FFC0C0C0")
            else:
                pass
            r += 1
        rangeobj = range(1, 10)
        for i in rangeobj:
            ws.cell(row=1, column=i).font = Font(bold=True)
            ws.cell(row=1, column=i).alignment = Alignment(horizontal="center")
        wb.save(self.fetch_filename_from_datetime(".xlsx"))